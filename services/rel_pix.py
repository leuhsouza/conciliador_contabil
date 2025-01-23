import pandas as pd
from openpyxl import load_workbook

# Caminho fixo para o arquivo de contas contábeis
contabil_file_path = r"Códigos.xlsx"

def criar_tabela_dinamica(df, output_file):
    # Código para gerar a tabela dinâmica permanece o mesmo
    df['Descrição'] = df.apply(lambda x: f"{int(x['Conta'])} {x['ContaDescricao']}|Contábil: {int(x['Contabil']) if pd.notna(x['Contabil']) else 747}", axis=1)
    df['Nome'] = df.apply(lambda x: f"{x['Documento']} {x['Matricula']} {x['Nome']}", axis=1)

    tabela_dinamica = pd.pivot_table(
        df,
        index=['Conta', 'Descrição', 'Situacao', 'Nome'],
        values=['Valor', 'Juros', 'ValorPag'],
        aggfunc='sum',
        margins=False
    )

    subtotais_descricao = tabela_dinamica.groupby(level=['Conta', 'Descrição']).sum()
    subtotais_descricao.index = pd.MultiIndex.from_tuples(
        [(conta, desc, 'Subtotal', '') for conta, desc in subtotais_descricao.index]
    )
    subtotais_descricao.index.names = tabela_dinamica.index.names

    subtotais_situacao = tabela_dinamica.groupby(level=['Conta', 'Descrição', 'Situacao']).sum()
    subtotais_situacao.index = pd.MultiIndex.from_tuples(
        [(conta, desc, sit, 'Subtotal') for conta, desc, sit in subtotais_situacao.index]
    )
    subtotais_situacao.index.names = tabela_dinamica.index.names

    tabela_completa = pd.concat([tabela_dinamica, subtotais_situacao, subtotais_descricao]).sort_index()

    totais_8888_9999 = pd.DataFrame({
        'Valor': [df.loc[df['Conta'] == '8888', 'Valor'].sum(), df.loc[df['Conta'] == '9999', 'Valor'].sum()],
        'Juros': [0, 0],
        'ValorPag': [df.loc[df['Conta'] == '8888', 'ValorPag'].sum(), df.loc[df['Conta'] == '9999', 'ValorPag'].sum()]
    }, index=pd.MultiIndex.from_tuples([
        (8888, 'Operações|Contábil: 8888', 'Subtotal', ''),
        (9999, 'Taxa|Contábil: 9999', 'Subtotal', '')
    ], names=['Conta', 'Descrição', 'Situacao', 'Nome']))

    tabela_completa = pd.concat([tabela_completa, totais_8888_9999])

    totais_gerais = pd.DataFrame(tabela_dinamica.sum()).T
    totais_gerais.index = pd.Index([('Total Geral', '', '', '')])
    totais_gerais.index.names = tabela_dinamica.index.names

    totais_gerais[['Valor', 'Juros', 'ValorPag']] += totais_8888_9999[['Valor', 'Juros', 'ValorPag']].sum()
    tabela_completa = pd.concat([tabela_completa, totais_gerais])

    tabela_completa = tabela_completa[['Valor', 'Juros', 'ValorPag']]

    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        tabela_completa.to_excel(writer, sheet_name='Relatório')

    print(f"Tabela dinâmica com subtotais salva com sucesso em: {output_file}")

import pandas as pd
from openpyxl import load_workbook

# Caminho fixo para o arquivo de contas contábeis
contabil_file_path = r"Códigos.xlsx"

def process_excel_pix(input_file, sheet_name, output_file):
    try:
        # Validar o caminho do arquivo e a planilha selecionada
        if not input_file:
            raise ValueError("Caminho do arquivo não fornecido.")
        if not sheet_name:
            raise ValueError("Nome da planilha não fornecido.")
        
        # Carregar o arquivo Excel
        workbook = load_workbook(input_file, data_only=True)
        sheets = workbook.sheetnames
        
        # Verificar se a planilha existe no arquivo
        if sheet_name not in sheets:
            raise ValueError(f"A planilha '{sheet_name}' não foi encontrada no arquivo '{input_file}'.")
        
        # Carregar a planilha selecionada
        df = pd.read_excel(input_file, sheet_name=sheet_name, header=None)
        
        # Validar se a planilha tem colunas suficientes
        if df.shape[1] < 9:
            raise ValueError("A planilha não possui o número mínimo de colunas para o processamento.")

        # Verifica a presença de 'operações' na coluna 7; se não encontrar, verifica na coluna B
        operacoes_rows = df.iloc[:, 7].str.contains('operações', case=False, na=False) | \
                        df.iloc[:, 1].str.contains('operações', case=False, na=False)

        # Verifica a presença de 'Taxa' na coluna 7; se não encontrar, verifica na coluna B
        taxa_rows = df.iloc[:, 7].str.contains('Taxa', case=False, na=False) | \
                    df.iloc[:, 1].str.contains('Taxa', case=False, na=False)

        # Busca os valores para 'operações'
        if operacoes_rows.any():
            operacoes_value = df.loc[operacoes_rows, df.columns[8]].values[0]  # Valor da coluna 8
            txt_operacoes = df.loc[operacoes_rows, df.columns[7]].values[0]  # Texto da coluna 7

            # Caso esteja ausente, tenta buscar na coluna C
            if pd.isna(operacoes_value):
                operacoes_value = df.loc[operacoes_rows, df.columns[2]].values[0]

        # Busca os valores para 'Taxa'
        if taxa_rows.any():
            taxa_value = df.loc[taxa_rows, df.columns[8]].values[0]  # Valor da coluna 8
            txt_taxa = df.loc[taxa_rows, df.columns[7]].values[0]  # Texto da coluna 7

            # Caso esteja ausente, tenta buscar na coluna C
            if pd.isna(taxa_value):
                taxa_value = df.loc[taxa_rows, df.columns[2]].values[0]

        # Define valores como None caso não sejam encontrados
        operacoes_value = operacoes_value if operacoes_rows.any() else None
        txt_operacoes = txt_operacoes if operacoes_rows.any() else None
        taxa_value = taxa_value if taxa_rows.any() else None
        txt_taxa = txt_taxa if taxa_rows.any() else None

        # Encontrar o índice da primeira linha de dados válidos
        first_value_index = df[df.iloc[:, 0].notna()].index[0]
        df.columns = df.iloc[first_value_index]
        df = df.iloc[first_value_index + 1:].reset_index(drop=True)

        if 'Documento' in df.columns:
            df['Documento'] = df['Documento'].astype(str).str.replace(r'^990000', '', regex=True)

        if 'TipoConta' in df.columns:
            df = df.drop('TipoConta', axis=1)

        # Adicionar valores de operações e taxas
        ultima_linha = df.dropna(how='all').index[-1]
        df.at[ultima_linha + 1, ('Valor', 'ValorPag')] = operacoes_value
        df.at[ultima_linha + 2, ('Valor', 'ValorPag')] = taxa_value
        df.at[ultima_linha + 1, ('ContaDescricao')] = txt_operacoes
        df.at[ultima_linha + 2, ('ContaDescricao')] = txt_taxa
        df.at[ultima_linha + 1, ('Conta')] = '8888'
        df.at[ultima_linha + 2, ('Conta')] = '9999'

        # Calcular juros
        df['Juros'] = df['ValorPag'] - df['Valor']
        df['Juros'] = df['Juros'].clip(lower=0)
        colunas = list(df.columns)
        colunas.insert(9, colunas.pop(colunas.index('Juros')))
        df = df[colunas]

        # Adicionar coluna 'Contabil'
        df.insert(3, 'Contabil', None)

        # Processar contas contábeis
        df_contabil = pd.read_excel(contabil_file_path, sheet_name="Planilha2", dtype={'Conta': str})
        df_contabil['Conta'] = df_contabil['Conta'].str.strip()
        df_contabil['Descricao'] = df_contabil['Descricao'].astype(str).str.strip()
        df_contabil['Chave'] = df_contabil['Conta'] + ' ' + df_contabil['Descricao']

        contas_especiais = ['0101', '0102', '0103', '0104', '0105', '0106']
        for i, row in df.iterrows():
            conta = str(row['Conta']).strip()
            situacao = str(row.get('Situacao', '')).strip()

            if conta in contas_especiais:
                chave = f"{conta} {situacao}"
                conta_contabil = df_contabil[df_contabil['Chave'] == chave]['Contabil'].values
            else:
                conta_contabil = df_contabil[df_contabil['Conta'] == conta]['Contabil'].values

            df.at[i, 'Contabil'] = conta_contabil[0] if conta_contabil.size > 0 else 747

        # Ordenar por conta
        df = df.sort_values("Conta")

        # Salvar o arquivo processado
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"Arquivo processado salvo em: {output_file}")
    
    except Exception as e:
        raise ValueError(f"Erro ao processar o arquivo: {e}")
    criar_tabela_dinamica(df, output_file)
